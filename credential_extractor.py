import re
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

def extract_credentials_from_xlsx(file_path, file_content=None):
    """Extract credentials from Excel files (.xlsx) - Real-world patterns"""
    if not XLSX_AVAILABLE:
        print("[-] openpyxl not available for XLSX processing. Install with: pip install openpyxl")
        return []
    
    credentials = []
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"  [*] Scanning sheet: {sheet_name}")
            
            # Convert sheet to list of rows for easier processing
            sheet_data = []
            max_row = min(sheet.max_row, 1000)  # Limit to first 1000 rows for performance
            max_col = min(sheet.max_column, 50)  # Limit to first 50 columns
            
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                sheet_data.append([str(cell).strip() if cell is not None else "" for cell in row])
            
            # Find header rows (usually in first 5 rows)
            header_mappings = []
            for header_row_idx in range(min(5, len(sheet_data))):
                header_row = sheet_data[header_row_idx]
                username_cols = []
                password_cols = []
                additional_cols = {}
                
                for col_idx, header in enumerate(header_row):
                    if not header:
                        continue
                    
                    header_lower = header.lower().strip()
                    
                    # Real-world username patterns
                    username_patterns = [
                        'username', 'user', 'login', 'account', 'user name', 'loginname',
                        'userid', 'user_id', 'user id', 'email', 'logon', 'signin',
                        'uname', 'uid', 'user_name', 'account_name', 'login_name',
                        'service_account', 'svc_account', 'admin', 'administrator',
                        'domain\\user', 'domain_user', 'ad_user', 'ldap_user'
                    ]
                    
                    # Real-world password patterns  
                    password_patterns = [
                        'password', 'pass', 'pwd', 'passwd', 'secret', 'key', 
                        'passphrase', 'pin', 'access_key', 'api_key', 'token',
                        'auth_token', 'authentication', 'credential', 'cred',
                        'service_password', 'admin_password', 'user_password',
                        'login_password', 'account_password', 'auth_key'
                    ]
                    
                    # Check for username columns
                    if any(pattern in header_lower for pattern in username_patterns):
                        username_cols.append(col_idx)
                        print(f"    [+] Found username column '{header}' at column {col_idx + 1}")
                    
                    # Check for password columns
                    elif any(pattern in header_lower for pattern in password_patterns):
                        password_cols.append(col_idx)
                        print(f"    [+] Found password column '{header}' at column {col_idx + 1}")
                    
                    # Additional useful columns
                    elif any(keyword in header_lower for keyword in ['server', 'host', 'system', 'application', 'service', 'database', 'url']):
                        additional_cols['target'] = col_idx
                    elif any(keyword in header_lower for keyword in ['description', 'notes', 'comment', 'purpose']):
                        additional_cols['description'] = col_idx
                    elif any(keyword in header_lower for keyword in ['role', 'privilege', 'access', 'permission']):
                        additional_cols['role'] = col_idx
                
                # If we found credential columns, record this mapping
                if username_cols and password_cols:
                    header_mappings.append({
                        'header_row': header_row_idx,
                        'username_cols': username_cols,
                        'password_cols': password_cols,
                        'additional_cols': additional_cols
                    })
                    print(f"    [+] Valid credential structure found in row {header_row_idx + 1}")
            
            # Extract credentials using found header mappings
            for mapping in header_mappings:
                header_row_idx = mapping['header_row']
                username_cols = mapping['username_cols']
                password_cols = mapping['password_cols']
                additional_cols = mapping['additional_cols']
                
                # Start scanning from the row after headers
                start_row = header_row_idx + 1
                
                for data_row_idx in range(start_row, len(sheet_data)):
                    data_row = sheet_data[data_row_idx]
                    
                    # Skip empty rows
                    if not any(cell.strip() for cell in data_row):
                        continue
                    
                    # Try all combinations of username/password columns
                    for username_col in username_cols:
                        for password_col in password_cols:
                            if username_col >= len(data_row) or password_col >= len(data_row):
                                continue
                            
                            username = data_row[username_col].strip()
                            password = data_row[password_col].strip()
                            
                            # Validate credentials
                            if (username and password and 
                                len(username) > 0 and len(password) > 0 and
                                username.lower() not in ['username', 'user', 'login', 'account', 'n/a', 'none', 'null', '', 'example', 'test'] and
                                password.lower() not in ['password', 'pass', 'pwd', 'secret', 'n/a', 'none', 'null', '', 'example', 'test', '*****', '****', 'xxxx']):
                                
                                # Get additional context
                                context_info = []
                                target = ""
                                description = ""
                                role = ""
                                
                                if 'target' in additional_cols and additional_cols['target'] < len(data_row):
                                    target = data_row[additional_cols['target']].strip()
                                    if target:
                                        context_info.append(f"Target: {target}")
                                
                                if 'description' in additional_cols and additional_cols['description'] < len(data_row):
                                    description = data_row[additional_cols['description']].strip()
                                    if description:
                                        context_info.append(f"Desc: {description}")
                                
                                if 'role' in additional_cols and additional_cols['role'] < len(data_row):
                                    role = data_row[additional_cols['role']].strip()
                                    if role:
                                        context_info.append(f"Role: {role}")
                                
                                context = f"Row {data_row_idx + 1}: {username} / {password[:20]}{'...' if len(password) > 20 else ''}"
                                if context_info:
                                    context += f" ({'; '.join(context_info)})"
                                
                                credentials.append({
                                    'type': 'credential_pair',
                                    'username': username,
                                    'password': password,
                                    'sheet': sheet_name,
                                    'row': data_row_idx + 1,
                                    'username_column': openpyxl.utils.get_column_letter(username_col + 1),
                                    'password_column': openpyxl.utils.get_column_letter(password_col + 1),
                                    'target': target,
                                    'description': description,
                                    'role': role,
                                    'context': context,
                                    'source': 'xlsx_structured_table'
                                })
            
            # Also check for single-cell colon-separated credentials (user:pass format)
            for row_idx, row in enumerate(sheet_data[:100]):  # Check first 100 rows only
                for col_idx, cell_value in enumerate(row):
                    if ':' in cell_value and len(cell_value.split(':')) == 2:
                        parts = cell_value.split(':')
                        potential_user = parts[0].strip()
                        potential_pass = parts[1].strip()
                        
                        # Validate it looks like credentials
                        if (len(potential_user) > 2 and len(potential_pass) > 2 and
                            not potential_user.replace('.', '').replace('-', '').isdigit() and  # Not IP address
                            potential_user.lower() not in ['http', 'https', 'ftp', 'ssh', 'tcp', 'udp'] and  # Not protocol
                            not any(word in cell_value.lower() for word in ['///', 'http://', 'https://', 'file://'])):  # Not URL
                            
                            credentials.append({
                                'type': 'credential_pair',
                                'username': potential_user,
                                'password': potential_pass,
                                'sheet': sheet_name,
                                'row': row_idx + 1,
                                'column': openpyxl.utils.get_column_letter(col_idx + 1),
                                'context': f"Cell {openpyxl.utils.get_column_letter(col_idx + 1)}{row_idx + 1}: {cell_value}",
                                'source': 'xlsx_colon_format'
                            })
        
        workbook.close()
        
        # Remove duplicates
        unique_creds = []
        seen = set()
        for cred in credentials:
            cred_key = (cred['username'].lower(), cred['password'], cred['sheet'])
            if cred_key not in seen:
                seen.add(cred_key)
                unique_creds.append(cred)
        
        print(f"  [+] Found {len(unique_creds)} unique credential pairs")
        return unique_creds
        
    except Exception as e:
        print(f"  [-] Error processing XLSX file: {str(e)}")
        return []


def parse_connection_string(conn_str):
    """Parse database connection strings for credentials"""
    credentials = []
    
    # Common connection string patterns
    patterns = [
        (r'User\s*Id\s*=\s*([^;]+)', 'username'),
        (r'Username\s*=\s*([^;]+)', 'username'),
        (r'Uid\s*=\s*([^;]+)', 'username'),
        (r'Password\s*=\s*([^;]+)', 'password'),
        (r'Pwd\s*=\s*([^;]+)', 'password'),
    ]
    
    found_creds = {}
    for pattern, field_type in patterns:
        match = re.search(pattern, conn_str, re.IGNORECASE)
        if match:
            found_creds[field_type] = match.group(1).strip('"\'')
    
    if 'username' in found_creds and 'password' in found_creds:
        credentials.append({
            'type': 'credential_pair',
            'username': found_creds['username'],
            'password': found_creds['password'],
            'source': 'connection_string'
        })
    
    return credentials

def is_key_file(file_path):
    """Check if file is a key/certificate file that shouldn't be parsed for credentials"""
    key_extensions = ['.key', '.pem', '.p12', '.pfx', '.crt', '.cer', '.der']
    file_ext = Path(file_path).suffix.lower()
    return file_ext in key_extensions or 'id_rsa' in Path(file_path).name.lower()

def extract_credentials_from_txt(file_path, file_content):
    """Extract credentials from text files - improved to reduce false positives"""
    credentials = []
    
    # More specific credential patterns to reduce false positives
    patterns = [
        # Direct username:password format (strict - no common config words)
        (r'^([a-zA-Z0-9_\-\.]+)\s*:\s*([^\s\n\r:]+)$', 'direct_colon_format'),
        
        # Explicit user/pass pairs (same line)
        (r'(?:user(?:name)?|login|account)\s*[:=]\s*([^\s\n\r]+).*?(?:pass(?:word)?|pwd)\s*[:=]\s*([^\s\n\r]+)', 'user_pass_pair'),
        (r'(?:pass(?:word)?|pwd)\s*[:=]\s*([^\s\n\r]+).*?(?:user(?:name)?|login|account)\s*[:=]\s*([^\s\n\r]+)', 'pass_user_pair'),
        
        # Command line patterns
        (r'(?i)-u\s+([^\s\n\r]+).*?-p\s+([^\s\n\r]+)', 'cmdline_up'),
        (r'(?i)-p\s+([^\s\n\r]+).*?-u\s+([^\s\n\r]+)', 'cmdline_pu'),
        (r'(?i)--user[=\s]+([^\s\n\r]+).*?--pass(?:word)?[=\s]+([^\s\n\r]+)', 'cmdline_long_up'),
        
        # URL with embedded credentials
        (r'(?i)://([^:]+):([^@]+)@', 'url_credentials'),
        (r'(?i)(ftp|ssh|mysql|postgresql|mongodb)://([^:]+):([^@]+)@', 'service_url_creds'),
        
        # Service account patterns (domain\user format)
        (r'(?i)([a-zA-Z0-9._-]+\\[a-zA-Z0-9._-]+)\s*[:=]\s*([^\s\n\r]+)', 'domain_user_pass'),
        (r'(?i)([a-zA-Z0-9._-]+@[a-zA-Z0-9._-]+)\s*[:=]\s*([^\s\n\r]+)', 'email_user_pass'),
        
        # JSON-style patterns
        (r'(?i)"(?:user|username|login)"\s*:\s*"([^"]+)".*?"(?:pass|password|pwd)"\s*:\s*"([^"]+)"', 'json_up'),
        
        # LDAP bind patterns
        (r'(?i)(bind_dn|binddn)\s*[:=]\s*([^\s\n\r]+)', 'ldap_bind_dn'),
        (r'(?i)(bind_pw|bindpw|bind_password)\s*[:=]\s*([^\s\n\r]+)', 'ldap_bind_pw'),
        
        # Specific credential patterns (not generic config)
        (r'(?i)\b(username|user|login|account)\s*[:=]\s*([^\s\n\r]+)', 'username_only'),
        (r'(?i)\b(password|pass|pwd|secret|api_key|auth_token)\s*[:=]\s*([^\s\n\r]+)', 'password_only'),
    ]
    
    # Words that indicate configuration, not credentials
    CONFIG_WORDS = {
        'host', 'hostname', 'server', 'port', 'database', 'db', 'table', 'schema',
        'timeout', 'ssl', 'tls', 'debug', 'verbose', 'log', 'path', 'dir', 'directory',
        'file', 'url', 'uri', 'protocol', 'version', 'encoding', 'charset', 'locale',
        'timezone', 'format', 'type', 'mode', 'level', 'size', 'limit', 'max', 'min',
        'enable', 'disable', 'true', 'false', 'yes', 'no', 'on', 'off'
    }
    
    lines = file_content.split('\n')
    single_creds = {}
    line_numbers = {}
    
    for i, line in enumerate(lines):
        line = line.strip()
        if not line or line.startswith('#') or line.startswith('//'):
            continue
            
        for pattern, pattern_type in patterns:
            matches = re.findall(pattern, line, re.IGNORECASE)
            if matches:
                for match in matches:
                    if pattern_type == 'direct_colon_format':
                        username, password = match
                        # Skip if username is a common config word
                        if (username.lower() not in CONFIG_WORDS and 
                            len(password) > 2 and 
                            not password.replace('.', '').replace(':', '').isdigit()):  # Skip IP addresses and ports
                            credentials.append({
                                'type': 'credential_pair',
                                'username': username,
                                'password': password,
                                'line': i + 1,
                                'context': line
                            })
                    
                    elif pattern_type in ['user_pass_pair', 'cmdline_up', 'cmdline_long_up', 'json_up']:
                        username, password = match
                        credentials.append({
                            'type': 'credential_pair',
                            'username': username,
                            'password': password,
                            'line': i + 1,
                            'context': line,
                            'source': pattern_type
                        })
                    
                    elif pattern_type in ['pass_user_pair', 'cmdline_pu']:
                        password, username = match
                        credentials.append({
                            'type': 'credential_pair',
                            'username': username,
                            'password': password,
                            'line': i + 1,
                            'context': line,
                            'source': pattern_type
                        })
                    
                    elif pattern_type in ['domain_user_pass', 'email_user_pass']:
                        username, password = match
                        credentials.append({
                            'type': 'credential_pair',
                            'username': username,
                            'password': password,
                            'line': i + 1,
                            'context': line,
                            'source': 'service_account_format'
                        })
                    
                    elif pattern_type in ['url_credentials', 'service_url_creds']:
                        if pattern_type == 'service_url_creds':
                            protocol, username, password = match
                            credentials.append({
                                'type': 'credential_pair',
                                'username': username,
                                'password': password,
                                'line': i + 1,
                                'context': line,
                                'source': f'{protocol}_url',
                                'protocol': protocol
                            })
                        else:
                            username, password = match
                            credentials.append({
                                'type': 'credential_pair',
                                'username': username,
                                'password': password,
                                'line': i + 1,
                                'context': line,
                                'source': 'url_embedded'
                            })
                    
                    elif pattern_type in ['username_only', 'ldap_bind_dn']:
                        field_type, value = match
                        field_key = field_type.lower()
                        
                        # Only store if it looks like actual credentials
                        if field_key in ['username', 'user', 'login', 'account'] and value.lower() not in CONFIG_WORDS:
                            single_creds['username'] = value
                            line_numbers['username'] = i + 1
                    
                    elif pattern_type in ['password_only', 'ldap_bind_pw']:
                        field_type, value = match
                        field_key = field_type.lower()
                        
                        if field_key in ['password', 'pass', 'pwd', 'secret', 'api_key', 'auth_token']:
                            # Skip obvious config values
                            if (not value.replace('.', '').replace(':', '').isdigit() and  # Not IP/port
                                value.lower() not in CONFIG_WORDS and
                                len(value) > 3):  # Reasonable password length
                                single_creds['password'] = value
                                line_numbers['password'] = i + 1
    
    # Try to pair username and password from separate lines
    if 'username' in single_creds and 'password' in single_creds:
        username_line = line_numbers['username']
        password_line = line_numbers['password']
        
        # Only pair if they're close (within 3 lines) and look legitimate
        if abs(username_line - password_line) <= 3:
            credentials.append({
                'type': 'credential_pair',
                'username': single_creds['username'],
                'password': single_creds['password'],
                'line': f"{username_line},{password_line}",
                'context': f"username from line {username_line}, password from line {password_line}",
                'source': 'multi_line_pairing'
            })
    
    return credentials
def analyze_file_for_credentials(file_path, share_name, remote_path):
    """Analyze a downloaded file for credentials based on its extension"""
    if not Path(file_path).exists():
        return []
    
    # Check if this is a key file first
    if is_key_file(file_path):
        print(f"  [+] Found key file: {Path(file_path).name}")
        return [{
            'type': 'key_file',
            'filename': Path(file_path).name,
            'file_type': 'cryptographic_key',
            'message': 'Cryptographic key or certificate file found',
            'file_path': file_path,
            'share_name': share_name,
            'remote_path': remote_path
        }]
    
    try:
        # For XLSX files, don't try to read as text
        file_ext = Path(file_path).suffix.lower()
        if file_ext in ['.xlsx', '.xlsm']:
            print(f"  [+] Extracting credentials from Excel file")
            credentials = extract_credentials_from_xlsx(file_path)
            
            # For XLSX, we need to read content differently to extract targets
            # Read as binary and convert what we can to text for target extraction
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                all_text = ""
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                all_text += str(cell) + " "
                workbook.close()
                potential_targets = extract_potential_targets_from_content(all_text)
            except:
                potential_targets = []
        else:
            # Original text-based extraction
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            if not content.strip():
                return []
            
            # Extract potential targets from content
            potential_targets = extract_potential_targets_from_content(content)
            
            if file_ext in ['.txt', '.log', '.conf', '.cfg', '.ini', '.sh', '.bash', '.zsh', '.properties', '.env', '.config']:
                print(f"  [+] Extracting credentials from text file")
                credentials = extract_credentials_from_txt(file_path, content)
            else:
                print(f"  [+] Extracting credentials from file (treating as text)")
                credentials = extract_credentials_from_txt(file_path, content)
    
    except Exception as e:
        print(f"  [-] Error reading {file_path}: {str(e)}")
        return []
    
    # Add file information and potential targets to each credential
    for cred in credentials:
        cred['file_path'] = file_path
        cred['share_name'] = share_name
        cred['remote_path'] = remote_path
        cred['potential_targets'] = potential_targets
    
    return credentials
def extract_potential_targets_from_content(file_content):
    """Extract potential target hosts/IPs from file content"""
    targets = set()
    
    # IP address pattern
    ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
    
    # Hostname patterns
    hostname_pattern = r'\b[a-zA-Z0-9]([a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?(\.[a-zA-Z0-9]([a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?)*\b'
    
    # URL patterns with embedded hosts
    url_patterns = [
        r'(?i)(?:ssh|ftp|ftps|sftp)://([^:/\s]+)',
        r'(?i)host\s*[:=]\s*([^\s\n\r;,]+)',
        r'(?i)server\s*[:=]\s*([^\s\n\r;,]+)',
        r'(?i)hostname\s*[:=]\s*([^\s\n\r;,]+)',
    ]
    
    # Find IPs
    for match in re.finditer(ip_pattern, file_content):
        ip = match.group()
        # Skip local/invalid IPs
        if not ip.startswith(('127.', '0.', '255.')):
            targets.add(ip)
    
    # Find URLs with hosts
    for pattern in url_patterns:
        for match in re.finditer(pattern, file_content):
            host = match.group(1).strip()
            if host and '.' in host and not host.replace('.', '').isdigit():
                targets.add(host)
    
    return list(targets)